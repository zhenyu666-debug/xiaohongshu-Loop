"""Fast reader for the Banking Fraud dataset xlsx.

Dataset: ``C:/Users/Hasee/Desktop/paysim_data/Banking Fruad dataset.xlsx``
- 9082 rows × 3925 cols (header row = F1..F3924, rows 1-9082 = data)
- Key features identified:
    col 0  : row index (1-based)
    col 13-19 : feature values (normalized 0-1, likely PCA or raw feature scores)
    col 3922  : isFraud target (~53% fraud in first 200 rows)
- Most other columns are sparse or 'NA'

The reader uses openpyxl read-only mode for memory efficiency and caches
a compact JSON snapshot so subsequent requests don't re-parse the xlsx.

Graph mapping for the D3 visualizer
-----------------------------------
Each row → one "Record" node in the graph.
Fraud records: red (isFraud=1)
Normal records: blue (isFraud=0)
Edges: connect records that share a similar value in a key feature column
       (threshold: |a - b| < 0.05).
"""

from __future__ import annotations

import json
import os
import sys
import time
from pathlib import Path
from typing import Any, Iterator

try:
    from openpyxl import load_workbook
except ImportError as e:
    raise SystemExit(f"openpyxl required: pip install openpyxl\n{e}")

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

XLSX_PATH = Path(os.environ.get(
    "BANKFRAUD_XLSX",
    r"C:\Users\Hasee\Desktop\paysim_data\Banking Fruad dataset.xlsx",
))

CACHE_JSON = Path(os.environ.get(
    "BANKFRAUD_CACHE",
    r"C:\Users\Hasee\Desktop\paysim_data\fraud500.json",
))

# Columns of interest (0-based within the xlsx)
# We only load these to keep memory low
COL_ROW_ID   = 0   # row index (1-based)
COL_ISFRAUD  = 3922  # isFraud target (column F3922)
# Feature columns we'll show in the stats panel
FEATURE_COLS = list(range(13, 20))  # cols 13-19 (7 features)
FEATURE_NAMES = [f"F{13+i}" for i in range(7)]


# ---------------------------------------------------------------------------
# Core reader
# ---------------------------------------------------------------------------

def _iter_rows_fast() -> Iterator[dict[str, Any]]:
    """Yield one dict per data row, only for columns we care about."""
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active

    # Skip header row, yield data rows
    row_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        # Safety cap — xlsx says 9082 data rows
        if row_num > 10000:
            break

        # Extract only what we need
        try:
            row_id   = row[COL_ROW_ID] if COL_ROW_ID < len(row) else row_num
            is_fraud = int(row[COL_ISFRAUD]) if COL_ISFRAUD < len(row) else 0
            features = []
            for c in FEATURE_COLS:
                v = row[c] if c < len(row) else None
                try:
                    features.append(float(v) if v is not None else 0.0)
                except (TypeError, ValueError):
                    features.append(0.0)
        except Exception:
            continue

        yield {
            "id": f"R{int(row_id) if isinstance(row_id, (int, float)) else row_num}",
            "is_fraud": 1 if is_fraud else 0,
            "features": features,
        }

    wb.close()


def load_full_data() -> list[dict[str, Any]]:
    """Load rows from JSON cache, normalizing field names."""
    if not CACHE_JSON.exists():
        raise FileNotFoundError(
            f"JSON cache not found: {CACHE_JSON}\n"
            "Run: python app/loader/bankfraud_loader.py to generate it from the xlsx."
        )
    print(f"[bankfraud] Loading from cache: {CACHE_JSON}")
    raw = json.loads(CACHE_JSON.read_text(encoding="utf-8"))
    rows = []
    for r in raw:
        # Normalize field names from xlsx/pandas column names
        is_fraud = int(r.get("F3922", r.get("is_fraud", 0)))
        row_id = r.get("Unnamed: 0", r.get("id", 0))
        features = [float(r.get(f"F{13+i}", 0.0) or 0.0) for i in range(7)]
        rows.append({
            "id": f"R{int(row_id)}",
            "is_fraud": is_fraud,
            "features": features,
        })
    print(f"[bankfraud] Loaded {len(rows)} rows")
    return rows


# ---------------------------------------------------------------------------
# Graph builder (for D3)
# ---------------------------------------------------------------------------

def build_graph(
    rows: list[dict[str, Any]],
    *,
    sample_size: int = 300,
    fraud_ratio: float = 0.5,
    seed: int = 42,
) -> dict[str, Any]:
    """Build graph nodes + edges from the banking fraud data.

    - sample_size: max total nodes
    - fraud_ratio: fraction of fraud nodes to include (balance the display)
    """
    import random
    rng = random.Random(seed)

    # Separate fraud / non-fraud
    fraud_rows = [r for r in rows if r["is_fraud"] == 1]
    normal_rows = [r for r in rows if r["is_fraud"] == 0]

    # Sample
    max_fraud = int(sample_size * fraud_ratio)
    max_normal = sample_size - max_fraud
    sampled = (
        rng.sample(fraud_rows, min(max_fraud, len(fraud_rows))) +
        rng.sample(normal_rows, min(max_normal, len(normal_rows)))
    )
    rng.shuffle(sampled)

    # Build nodes
    nodes = []
    for r in sampled:
        nodes.append({
            "id": r["id"],
            "label": f"R{r['id'][1:]}",
            "type": "fraud" if r["is_fraud"] else "normal",
            "is_fraud": r["is_fraud"],
            "color": "#ff5d6c" if r["is_fraud"] else "#6ad1ff",
            "radius": 9 if r["is_fraud"] else 7,
            "features": r["features"],
        })

    # Build edges: connect records with similar feature values
    # Threshold: |a - b| < 0.05 for at least 2 features
    THRESHOLD = 0.05
    MIN_COMMON = 2

    edges = []
    edge_set: set[tuple[str, str]] = set()
    n = len(nodes)
    for i in range(n):
        for j in range(i + 1, n):
            if len(edges) >= 200:  # cap edges for display
                break
            ni, nj = nodes[i], nodes[j]
            fi, fj = ni["features"], nj["features"]
            common = sum(1 for a, b in zip(fi, fj) if abs(a - b) < THRESHOLD)
            if common >= MIN_COMMON:
                key = (ni["id"], nj["id"])
                if key not in edge_set:
                    edge_set.add(key)
                    is_same = ni["is_fraud"] == nj["is_fraud"]
                    edges.append({
                        "source": ni["id"],
                        "target": nj["id"],
                        "type": "SIMILAR",
                        "color": "#ff5d6c55" if not is_same else "#6ad1ff44",
                        "amount": common,
                    })
            if len(edges) >= 200:
                break

    return {"nodes": nodes, "edges": edges}


# ---------------------------------------------------------------------------
# Statistics
# ---------------------------------------------------------------------------

def compute_stats(rows: list[dict[str, Any]]) -> dict[str, Any]:
    fraud = [r for r in rows if r["is_fraud"] == 1]
    normal = [r for r in rows if r["is_fraud"] == 0]
    total = len(rows)

    # Feature distributions for fraud vs normal
    fraud_feats = [r["features"] for r in fraud]
    normal_feats = [r["features"] for r in normal]

    def mean_feat(feats: list[list[float]]) -> list[float]:
        if not feats:
            return [0.0] * 7
        return [sum(f[i] for f in feats) / len(feats) for i in range(7)]

    fraud_mean = mean_feat(fraud_feats)
    normal_mean = mean_feat(normal_feats)

    return {
        "total_count": total,
        "fraud_count": len(fraud),
        "normal_count": len(normal),
        "fraud_rate": round(len(fraud) / max(1, total) * 100, 2),
        "feature_names": FEATURE_NAMES,
        "fraud_mean": [round(v, 4) for v in fraud_mean],
        "normal_mean": [round(v, 4) for v in normal_mean],
    }


# ---------------------------------------------------------------------------
# API serializer
# ---------------------------------------------------------------------------

def build_api_response(
    rows: list[dict[str, Any]] | None = None,
    *,
    sample_size: int = 300,
    fraud_ratio: float = 0.5,
) -> dict[str, Any]:
    """Build the full API response payload."""
    rows = rows or load_full_data()
    stats = compute_stats(rows)
    graph = build_graph(rows, sample_size=sample_size, fraud_ratio=fraud_ratio)

    return {
        "ok": True,
        "source": str(XLSX_PATH),
        "total_rows": len(rows),
        "stats": stats,
        "nodes": graph["nodes"],
        "edges": graph["edges"],
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    import argparse
    global XLSX_PATH
    parser = argparse.ArgumentParser(description="Banking fraud dataset loader")
    parser.add_argument("--xlsx", default=str(XLSX_PATH), help="Path to xlsx file")
    parser.add_argument("--sample", type=int, default=300, help="Graph sample size")
    parser.add_argument("--fraud-ratio", type=float, default=0.5, help="Fraud node ratio")
    args = parser.parse_args(argv)

    XLSX_PATH = Path(args.xlsx)

    resp = build_api_response(sample_size=args.sample, fraud_ratio=args.fraud_ratio)
    print(json.dumps(resp["stats"], indent=2))
    print(f"\nGraph: {len(resp['nodes'])} nodes, {len(resp['edges'])} edges")
    return 0


if __name__ == "__main__":
    sys.exit(main())
